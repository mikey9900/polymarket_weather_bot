"""Human-readable Excel export for analysis bundles."""

from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

TITLE_FILL = PatternFill(start_color="0B1F3A", end_color="0B1F3A", fill_type="solid")
SUBTITLE_FILL = PatternFill(start_color="DCE6F2", end_color="DCE6F2", fill_type="solid")
SECTION_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
LABEL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
ALT_FILL = PatternFill(start_color="F7FBFF", end_color="F7FBFF", fill_type="solid")
OPEN_FILL = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
RESOLVED_FILL = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
CLOSED_FILL = PatternFill(start_color="F3F3F3", end_color="F3F3F3", fill_type="solid")
GOOD_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
BAD_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
WARN_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
YES_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
NO_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
CONFIRMED_FILL = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
PROBABLE_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
SINGLE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

TITLE_FONT = Font(color="FFFFFF", bold=True, size=15)
SUBTITLE_FONT = Font(color="1F1F1F", italic=True, size=10)
SECTION_FONT = Font(color="FFFFFF", bold=True, size=11)
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
LABEL_FONT = Font(color="1F1F1F", bold=True, size=10)
GOOD_FONT = Font(color="006100", bold=True)
BAD_FONT = Font(color="9C0006", bold=True)
MUTED_FONT = Font(color="666666")

THIN_SIDE = Side(style="thin", color="B8C7DA")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)


def build_analysis_report(
    *,
    output_path: str | Path,
    label: str,
    created_at: datetime,
    snapshot: dict[str, Any],
    tracker,
    runtime,
) -> dict[str, Any]:
    """Create a styled workbook for the current bot state."""

    workbook = Workbook()
    stale_after = getattr(getattr(runtime, "config", None), "paper", None)
    stale_after = getattr(stale_after, "mark_stale_after_seconds", None)

    open_positions = tracker.get_dashboard_paper_positions(
        limit=250,
        status="open",
        mark_stale_after_seconds=stale_after,
    )
    recent_positions = tracker.get_dashboard_paper_positions(
        limit=250,
        mark_stale_after_seconds=stale_after,
    )
    recent_outcomes = tracker.get_dashboard_paper_positions(
        limit=250,
        statuses=("closed", "resolved"),
        mark_stale_after_seconds=stale_after,
    )
    recent_signals = tracker.get_recent_signals(limit=250)
    recent_resolutions = tracker.get_recent_resolutions(limit=250)
    recent_operator_actions = tracker.get_recent_operator_actions(limit=250)

    overview = workbook.active
    overview.title = "Overview"
    _build_overview_sheet(
        overview,
        label=label,
        created_at=created_at,
        snapshot=snapshot,
        open_positions=open_positions,
        recent_outcomes=recent_outcomes,
        recent_signals=recent_signals,
        recent_operator_actions=recent_operator_actions,
    )
    _build_positions_sheet(workbook.create_sheet("Open Trades"), open_positions, title="Open Trades", status_default="open")
    _build_outcomes_sheet(workbook.create_sheet("Recent Outcomes"), recent_outcomes)
    _build_positions_sheet(workbook.create_sheet("Position Ledger"), recent_positions, title="Position Ledger", status_default="")
    _build_signals_sheet(workbook.create_sheet("Recent Signals"), recent_signals)
    _build_resolutions_sheet(workbook.create_sheet("Resolutions"), recent_resolutions)
    _build_operator_sheet(workbook.create_sheet("Operator Log"), recent_operator_actions)

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return {
        "report_path": str(output),
        "sheet_count": len(workbook.sheetnames),
        "open_position_count": len(open_positions),
        "recent_outcome_count": len(recent_outcomes),
        "recent_signal_count": len(recent_signals),
    }


def _build_overview_sheet(
    ws,
    *,
    label: str,
    created_at: datetime,
    snapshot: dict[str, Any],
    open_positions: list[dict[str, Any]],
    recent_outcomes: list[dict[str, Any]],
    recent_signals: list[dict[str, Any]],
    recent_operator_actions: list[dict[str, Any]],
) -> None:
    ws.merge_cells("A1:H1")
    ws["A1"] = f"{label} Analysis Report"
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:H2")
    ws["A2"] = (
        f"Generated {created_at.astimezone(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')} | "
        "Human-readable export for paper-trade review and Dropbox sharing"
    )
    ws["A2"].fill = SUBTITLE_FILL
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].alignment = Alignment(horizontal="center")

    paper = (((snapshot or {}).get("summary") or {}).get("paper") or {})
    controls = (snapshot or {}).get("controls") or {}
    runtime = (snapshot or {}).get("runtime") or {}
    exports = (snapshot or {}).get("exports") or {}
    signal_summary = (snapshot or {}).get("signal_summary_24h") or {}

    _write_metric_block(
        ws,
        start_col=1,
        start_row=4,
        title="Paper Performance",
        items=[
            ("Initial Capital", paper.get("initial"), "currency_plain"),
            ("Cash Balance", paper.get("balance"), "currency_plain"),
            ("Current Equity", paper.get("equity"), "currency_plain"),
            ("Total P/L", paper.get("pnl"), "currency"),
            ("Wins", paper.get("wins"), "int"),
            ("Losses", paper.get("losses"), "int"),
            ("Win Rate", _percent_value(paper.get("win_rate")), "percent"),
            ("Open Positions", paper.get("open_positions"), "int"),
        ],
    )
    _write_metric_block(
        ws,
        start_col=4,
        start_row=4,
        title="Runtime Status",
        items=[
            ("State", str(controls.get("state") or "unknown").upper(), "text"),
            ("Temp Scope", controls.get("temperature_market_scope"), "text"),
            ("Scan Queue", controls.get("scan_queue_depth"), "int"),
            ("Worker Healthy", _bool_label(controls.get("scan_worker_healthy")), "text"),
            ("Last Temp Scan", runtime.get("last_temperature_scan_status"), "text"),
            ("Last Review", controls.get("last_open_position_review_status"), "text"),
            ("Temp Signals 24h", signal_summary.get("temperature"), "int"),
            ("Rain Signals 24h", signal_summary.get("precipitation"), "int"),
        ],
    )
    _write_metric_block(
        ws,
        start_col=7,
        start_row=4,
        title="Export Status",
        items=[
            ("Bundle", exports.get("last_analysis_bundle_path") or exports.get("latest_analysis_bundle_path"), "text"),
            ("Excel Report", exports.get("latest_analysis_report_path"), "text"),
            ("Dropbox Ready", _bool_label(exports.get("analysis_dropbox_enabled")), "text"),
            ("Cloud Link", exports.get("last_analysis_report_dropbox_url") or exports.get("last_analysis_bundle_dropbox_url"), "text"),
            ("Mirror JSON", exports.get("dashboard_state_path"), "text"),
            ("Recent Signals", len(recent_signals), "int"),
            ("Open Cards", len(open_positions), "int"),
            ("Closed/Resolved", len(recent_outcomes), "int"),
        ],
    )

    preview_row = 15
    ws.merge_cells(start_row=preview_row, start_column=1, end_row=preview_row, end_column=8)
    ws.cell(row=preview_row, column=1, value="Quick Read").fill = SECTION_FILL
    ws.cell(row=preview_row, column=1).font = SECTION_FONT
    ws.cell(row=preview_row, column=1).alignment = Alignment(horizontal="center")
    quick_lines = [
        f"Open positions: {len(open_positions)} | Recent outcomes: {len(recent_outcomes)} | Recent signals: {len(recent_signals)}",
        f"Last operator action: {recent_operator_actions[0]['action'] if recent_operator_actions else 'none'}",
        f"Latest cloud report: {exports.get('last_analysis_report_dropbox_url') or 'not uploaded yet'}",
    ]
    for idx, line in enumerate(quick_lines, start=preview_row + 1):
        ws.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=8)
        cell = ws.cell(row=idx, column=1, value=line)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if idx % 2 == 0:
            cell.fill = ALT_FILL

    for column, width in {"A": 18, "B": 18, "C": 18, "D": 18, "E": 18, "F": 18, "G": 18, "H": 24}.items():
        ws.column_dimensions[column].width = width


def _build_positions_sheet(ws, rows: list[dict[str, Any]], *, title: str, status_default: str) -> None:
    _sheet_title(ws, title, subtitle=f"{len(rows)} rows")
    columns = [
        ("Event", "event_title", "text"),
        ("City", "city_slug", "text"),
        ("Type", "market_type", "text"),
        ("Date", "event_date", "text"),
        ("Direction", "direction", "text"),
        ("Target", "target_label", "text"),
        ("Status", "status", "text"),
        ("Stake $", "cost", "currency"),
        ("Shares", "shares", "number"),
        ("Entry %", "entry_price", "percent"),
        ("Mark %", "market_probability", "percent"),
        ("Model %", "outcome_probability", "percent"),
        ("Mark P/L $", "mark_to_market_pnl", "currency"),
        ("Model P/L $", "expected_value_pnl", "currency"),
        ("Edge %", "mark_edge_abs", "percent"),
        ("Confidence", "confidence", "text"),
        ("Sources", "source_count", "int"),
        ("Held", "holding_seconds", "duration"),
        ("Review Age", "mark_age_seconds", "duration"),
        ("Review Note", "mark_reason", "text"),
        ("Market Slug", "market_slug", "text"),
    ]
    _write_table(ws, start_row=4, columns=columns, rows=rows)
    _style_position_rows(ws, rows, start_row=5, columns=columns, status_default=status_default)


def _build_signals_sheet(ws, rows: list[dict[str, Any]]) -> None:
    _sheet_title(ws, "Recent Signals", subtitle=f"{len(rows)} rows")
    columns = [
        ("Event", "event_title", "text"),
        ("City", "city_slug", "text"),
        ("Type", "market_type", "text"),
        ("Date", "event_date", "text"),
        ("Direction", "direction", "text"),
        ("Label", "label", "text"),
        ("Market %", "market_prob", "percent"),
        ("Forecast %", "forecast_prob", "percent"),
        ("Edge %", "edge_abs", "percent"),
        ("Confidence", "confidence", "text"),
        ("Sources", "source_count", "int"),
        ("Liquidity", "liquidity", "currency"),
        ("Score", "score", "number"),
        ("Created", "created_at", "datetime"),
        ("Market Slug", "market_slug", "text"),
    ]
    _write_table(ws, start_row=4, columns=columns, rows=rows)
    for idx, row in enumerate(rows, start=5):
        _fill_if_text(ws.cell(row=idx, column=5), row.get("direction"))
        _fill_confidence(ws.cell(row=idx, column=10), row.get("confidence"))


def _build_outcomes_sheet(ws, rows: list[dict[str, Any]]) -> None:
    _sheet_title(ws, "Recent Outcomes", subtitle=f"{len(rows)} rows")
    columns = [
        ("Event", "event_title", "text"),
        ("City", "city_slug", "text"),
        ("Type", "market_type", "text"),
        ("Date", "event_date", "text"),
        ("Direction", "direction", "text"),
        ("Target", "target_label", "text"),
        ("Status", "status", "text"),
        ("Outcome", "resolution", "text"),
        ("Stake $", "cost", "currency_plain"),
        ("Exit %", "exit_reference_price", "percent"),
        ("Realized P/L $", "realized_pnl", "currency"),
        ("Model P/L $", "expected_value_pnl", "currency"),
        ("Exit Reason", "exit_reason", "text"),
        ("Resolved", "resolved_at", "datetime"),
        ("Market Slug", "market_slug", "text"),
    ]
    _write_table(ws, start_row=4, columns=columns, rows=rows)
    for idx, row in enumerate(rows, start=5):
        _fill_if_text(ws.cell(row=idx, column=5), row.get("direction"))
        _fill_status(ws.cell(row=idx, column=7), row.get("status"))
        _fill_if_text(ws.cell(row=idx, column=8), row.get("resolution"))
        _pnl_style(ws.cell(row=idx, column=11))
        _pnl_style(ws.cell(row=idx, column=12))


def _build_resolutions_sheet(ws, rows: list[dict[str, Any]]) -> None:
    _sheet_title(ws, "Resolutions", subtitle=f"{len(rows)} rows")
    columns = [
        ("When", "resolved_at", "datetime"),
        ("Market Slug", "market_slug", "text"),
        ("Status", "status", "text"),
        ("Outcome", "outcome_label", "text"),
        ("Positions", "resolved_positions", "int"),
        ("Payout $", "total_payout", "currency"),
        ("Realized P/L $", "total_realized_pnl", "currency"),
    ]
    _write_table(ws, start_row=4, columns=columns, rows=rows)
    for idx, row in enumerate(rows, start=5):
        _fill_status(ws.cell(row=idx, column=3), row.get("status"))
        _pnl_style(ws.cell(row=idx, column=7))


def _build_operator_sheet(ws, rows: list[dict[str, Any]]) -> None:
    _sheet_title(ws, "Operator Log", subtitle=f"{len(rows)} rows")
    shaped = []
    for row in rows:
        payload = row.get("payload") or {}
        shaped.append(
            {
                "created_at": row.get("created_at"),
                "action": row.get("action"),
                "ok": payload.get("ok"),
                "status": payload.get("status"),
                "message": payload.get("message"),
                "payload": json.dumps(payload, sort_keys=True),
            }
        )
    columns = [
        ("When", "created_at", "datetime"),
        ("Action", "action", "text"),
        ("OK", "ok", "text"),
        ("Status", "status", "int"),
        ("Message", "message", "text"),
        ("Payload", "payload", "text"),
    ]
    _write_table(ws, start_row=4, columns=columns, rows=shaped)
    for idx, row in enumerate(shaped, start=5):
        ok_value = str(row.get("ok") or "").strip().lower()
        if ok_value == "true":
            ws.cell(row=idx, column=3).fill = GOOD_FILL
            ws.cell(row=idx, column=3).font = GOOD_FONT
        elif ok_value == "false":
            ws.cell(row=idx, column=3).fill = BAD_FILL
            ws.cell(row=idx, column=3).font = BAD_FONT


def _sheet_title(ws, title: str, *, subtitle: str = "") -> None:
    ws.merge_cells("A1:J1")
    ws["A1"] = title
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:J2")
    ws["A2"] = subtitle
    ws["A2"].fill = SUBTITLE_FILL
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].alignment = Alignment(horizontal="center")


def _write_metric_block(ws, *, start_col: int, start_row: int, title: str, items: list[tuple[str, Any, str]]) -> None:
    ws.merge_cells(
        start_row=start_row,
        start_column=start_col,
        end_row=start_row,
        end_column=start_col + 1,
    )
    header = ws.cell(row=start_row, column=start_col, value=title)
    header.fill = SECTION_FILL
    header.font = SECTION_FONT
    header.alignment = Alignment(horizontal="center")
    for offset, (label, value, kind) in enumerate(items, start=1):
        label_cell = ws.cell(row=start_row + offset, column=start_col, value=label)
        label_cell.fill = LABEL_FILL
        label_cell.font = LABEL_FONT
        label_cell.border = THIN_BORDER
        value_cell = ws.cell(row=start_row + offset, column=start_col + 1, value=_coerce_value(value, kind))
        value_cell.border = THIN_BORDER
        value_cell.alignment = Alignment(horizontal="right" if kind in {"currency", "percent", "number", "int"} else "left")
        _apply_number_format(value_cell, kind)
        if kind == "currency":
            _pnl_style(value_cell)
        if kind == "text":
            if str(value or "").strip().lower() in {"yes", "no"}:
                _fill_if_text(value_cell, value)
            elif str(value or "").strip().lower() in {"confirmed", "probable", "single_source", "weatherapi_only"}:
                _fill_confidence(value_cell, value)


def _write_table(
    ws,
    *,
    start_row: int,
    columns: list[tuple[str, str, str]],
    rows: list[dict[str, Any]],
) -> None:
    header_row = start_row
    for col_idx, (header, _, _) in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    for row_idx, row in enumerate(rows, start=header_row + 1):
        if row_idx % 2 == 0:
            for col_idx in range(1, len(columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = ALT_FILL
        for col_idx, (_, key, kind) in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_coerce_value(row.get(key), kind))
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                vertical="top",
                horizontal="right" if kind in {"currency", "percent", "number", "int"} else "left",
                wrap_text=kind in {"text"},
            )
            _apply_number_format(cell, kind)

    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(columns))}{max(header_row, header_row + len(rows))}"
    _auto_width(ws, column_count=len(columns))


def _style_position_rows(
    ws,
    rows: list[dict[str, Any]],
    *,
    start_row: int,
    columns: list[tuple[str, str, str]],
    status_default: str,
) -> None:
    column_lookup = {name: idx + 1 for idx, (name, _, _) in enumerate(columns)}
    for row_idx, row in enumerate(rows, start=start_row):
        status_value = str(row.get("status") or status_default or "").strip().lower()
        row_fill = OPEN_FILL if status_value == "open" else RESOLVED_FILL if status_value == "resolved" else CLOSED_FILL
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.fill.patternType == "solid" and cell.fill.fgColor.rgb not in {None, "00000000"}:
                continue
            cell.fill = row_fill if row_idx % 2 else cell.fill
        _fill_if_text(ws.cell(row=row_idx, column=column_lookup["Direction"]), row.get("direction"))
        _fill_status(ws.cell(row=row_idx, column=column_lookup["Status"]), row.get("status") or status_default)
        _fill_confidence(ws.cell(row=row_idx, column=column_lookup["Confidence"]), row.get("confidence"))
        _pnl_style(ws.cell(row=row_idx, column=column_lookup["Mark P/L $"]))
        _pnl_style(ws.cell(row=row_idx, column=column_lookup["Model P/L $"]))


def _auto_width(ws, *, column_count: int, cap: int = 34) -> None:
    for column_index in range(1, column_count + 1):
        column_letter = get_column_letter(column_index)
        max_len = 0
        for cell in ws[column_letter]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_len + 3, cap)


def _coerce_value(value: Any, kind: str) -> Any:
    if kind == "currency":
        return _as_float(value)
    if kind == "currency_plain":
        return _as_float(value)
    if kind == "percent":
        return _percent_value(value)
    if kind == "number":
        return _as_float(value)
    if kind == "int":
        try:
            return int(float(value or 0))
        except (TypeError, ValueError):
            return 0
    if kind == "datetime":
        return _format_datetime(value)
    if kind == "duration":
        return _format_duration(value)
    return str(value or "")


def _apply_number_format(cell, kind: str) -> None:
    if kind in {"currency", "currency_plain"}:
        cell.number_format = '$#,##0.00;[Red]-$#,##0.00'
    elif kind == "percent":
        cell.number_format = "0.0%"
    elif kind == "number":
        cell.number_format = "0.00"
    elif kind == "int":
        cell.number_format = "0"
    elif kind in {"text", "datetime", "duration"}:
        if kind == "text" and len(str(cell.value or "")) > 34:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _pnl_style(cell) -> None:
    try:
        value = float(cell.value or 0.0)
    except (TypeError, ValueError):
        return
    if value > 0:
        cell.fill = GOOD_FILL
        cell.font = GOOD_FONT
    elif value < 0:
        cell.fill = BAD_FILL
        cell.font = BAD_FONT


def _fill_if_text(cell, value: Any) -> None:
    text = str(value or "").strip().upper()
    if text == "YES":
        cell.fill = YES_FILL
    elif text == "NO":
        cell.fill = NO_FILL


def _fill_status(cell, value: Any) -> None:
    status = str(value or "").strip().lower()
    if status == "open":
        cell.fill = OPEN_FILL
    elif status == "resolved":
        cell.fill = RESOLVED_FILL
    elif status == "closed":
        cell.fill = CLOSED_FILL


def _fill_confidence(cell, value: Any) -> None:
    confidence = str(value or "").strip().lower()
    if confidence == "confirmed":
        cell.fill = CONFIRMED_FILL
    elif confidence in {"probable", "supported"}:
        cell.fill = PROBABLE_FILL
    elif confidence:
        cell.fill = SINGLE_FILL


def _format_datetime(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError:
        return text
    return parsed.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")


def _format_duration(value: Any) -> str:
    seconds = _as_float(value)
    if seconds is None:
        return ""
    remaining = int(round(seconds))
    sign = "-" if remaining < 0 else ""
    remaining = abs(remaining)
    days, remaining = divmod(remaining, 86400)
    hours, remaining = divmod(remaining, 3600)
    minutes, secs = divmod(remaining, 60)
    parts = []
    if days:
        parts.append(f"{days}d")
    if hours or days:
        parts.append(f"{hours}h")
    if minutes or hours or days:
        parts.append(f"{minutes}m")
    parts.append(f"{secs}s")
    return sign + " ".join(parts[:4])


def _percent_value(value: Any) -> float | None:
    numeric = _as_float(value)
    if numeric is None:
        return None
    if numeric > 1.0:
        return numeric / 100.0
    return numeric


def _as_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _bool_label(value: Any) -> str:
    if value is True:
        return "YES"
    if value is False:
        return "NO"
    text = str(value or "").strip()
    if not text:
        return ""
    return text.upper()
